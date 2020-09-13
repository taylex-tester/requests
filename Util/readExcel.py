import json
import os
import openpyxl

from Util.basePath import base_path


class ReadExcel(object):
    @staticmethod
    def load(filename="testCase.xlsx"):
        excel_file = os.path.join(base_path, "Config", filename)
        workbook = openpyxl.load_workbook(excel_file)
        return workbook

    def get_sheet(self, index=0):
        if index > len(self.load().sheetnames):
            return Exception("sheetIndex out of range!!")
        sheet = self.load().worksheets[index]
        return sheet

    def get_rows(self):
        return self.get_sheet().max_row

    def get_row_list(self, row_index):
        if 0 < row_index <= self.get_rows():
            row_gen = self.get_sheet()[row_index]
            # strip()确保value没有空格
            row_list = [a.value for a in row_gen]
            return row_list
        else:
            return Exception("row_index out of range!!")

    def get_col_list(self, col_index):
        if 0 < col_index <= self.get_sheet().max_column:
            gen = self.get_sheet().rows
            col_list = [i[col_index - 1].value for i in gen]
            return col_list
        else:
            return Exception("col_index out of range!!")

    def excel_list(self):
        excel_list = []
        for index in range(1, self.get_rows()):
            row_list = self.get_row_list(index + 1)
            excel_list.append(row_list)
        return excel_list

    def write(self, row, col, value, filename="testCase.xlsx"):
        excel_file = os.path.join(base_path, "Config", filename)
        workbook = self.load()
        to_write = workbook.active
        to_write.cell(row=row, column=col, value=value)
        workbook.save(excel_file)

    def get_row_number(self, caseId):
        num = 1
        for i in self.get_col_list(1):
            if i == caseId:
                return num
            num += 1
        return num


readExcel = ReadExcel()
if __name__ == '__main__':
    a = ReadExcel()
    # print(a.excel_list()[0][3])
    # b = a.excel_list()[0][3]
    # print(type(json.loads(b)))
    print(a.get_row_number("demo_001"))
